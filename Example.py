import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook

#Opening the workbook
book = load_workbook('Example.xlsx')
sheet = book.active

c1 = sheet.cell(row = 1, column = 1)

# writing values to cells
c1.value = "Raza"

c2 = sheet.cell(row= 1 , column = 2)
c2.value = "6"

c3 = sheet['A2']
c3.value = "Zohair"

# B2 means column = 2 & row = 2.
c4 = sheet['B2']
c4.value = "1"

book.save("Example.xlsx")
